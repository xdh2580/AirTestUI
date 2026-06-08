"""
数据驱动加载器模块
支持 YAML / Excel / JSON 格式的测试数据加载
"""
import json
from pathlib import Path

import yaml
from config.settings import ROOT_DIR
from utils.logger import get_logger

log = get_logger("DataLoader")

# 测试数据目录
DATA_DIR = ROOT_DIR / "data"


def load_yaml(filepath: str) -> dict | list:
    """
    加载 YAML 测试数据文件

    Args:
        filepath: 相对于 data/ 目录的文件路径，或绝对路径

    Returns:
        解析后的数据

    用法:
        data = load_yaml("test_data.yaml")
        data = load_yaml("subdir/login_data.yaml")
    """
    path = _resolve_path(filepath)
    with open(path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    log.debug(f"加载 YAML 数据: {path} ({len(data) if data else 0} 条)")
    return data or {}


def load_json(filepath: str) -> dict | list:
    """加载 JSON 测试数据文件"""
    path = _resolve_path(filepath)
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    log.debug(f"加载 JSON 数据: {path}")
    return data


def load_excel(filepath: str, sheet_name: str = None) -> list[dict]:
    """
    加载 Excel 测试数据文件

    Args:
        filepath: 文件路径
        sheet_name: 工作表名称，为空则取第一个

    Returns:
        字典列表，每行数据为一个字典（首行为键名）
    """
    from openpyxl import load_workbook

    path = _resolve_path(filepath)
    wb = load_workbook(path, read_only=True, data_only=True)

    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return []

    headers = rows[0]
    data = []
    for row in rows[1:]:
        row_data = {}
        for key, value in zip(headers, row):
            if key is not None:
                row_data[str(key)] = value
        if row_data:
            data.append(row_data)

    wb.close()
    log.debug(f"加载 Excel 数据: {path}, 工作表: {sheet_name or '默认'}, {len(data)} 行")
    return data


def parametrize_data(filepath: str, key: str = None) -> list:
    """
    加载数据并格式化为 pytest.mark.parametrize 可用的参数列表

    Args:
        filepath: 数据文件路径
        key: YAML 中要提取的顶层键名（仅 YAML 有效）

    Returns:
        参数列表

    用法:
        @pytest.mark.parametrize("case", parametrize_data("login.yaml", "login_cases"))
        def test_login(case):
            ...
    """
    path = Path(filepath)
    suffix = path.suffix.lower()

    if suffix in (".yaml", ".yml"):
        data = load_yaml(filepath)
        if key and isinstance(data, dict):
            data = data.get(key, [])
    elif suffix == ".json":
        data = load_json(filepath)
        if key and isinstance(data, dict):
            data = data.get(key, [])
    elif suffix in (".xlsx", ".xls"):
        data = load_excel(filepath)
    else:
        raise ValueError(f"不支持的数据文件格式: {suffix}")

    if isinstance(data, list):
        return data
    return [data]


def _resolve_path(filepath: str) -> Path:
    """解析文件路径，支持相对路径和绝对路径"""
    path = Path(filepath)
    if path.is_absolute():
        return path
    return DATA_DIR / filepath
